from requests import get as get_requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
data = [
    {
        "problem": "egg",
        "c": "No submission",
        "cpp": "http://moss.stanford.edu/results/7/7368257102529",
        "java": "No submission",
        "python": "No submission",
        "pascal": "No submission",
        "moss": "100"
    },
    {
        "problem": "bai4",
        "c": "No submission",
        "cpp": "http://moss.stanford.edu/results/7/6270151709645",
        "java": "No submission",
        "python": "No submission",
        "pascal": "No submission",
        "moss": "100"
    },
    {
        "problem": "bai5",
        "c": "No submission",
        "cpp": "http://moss.stanford.edu/results/7/4693195625181",
        "java": "No submission",
        "python": "No submission",
        "pascal": "No submission",
        "moss": "100"
    },
    {
        "problem": "bai6",
        "c": "No submission",
        "cpp": "http://moss.stanford.edu/results/1/1978525613407",
        "java": "No submission",
        "python": "No submission",
        "pascal": "No submission",
        "moss": "100"
    },
    {
        "problem": "bank",
        "c": "No submission",
        "cpp": "http://moss.stanford.edu/results/9/7447474171380",
        "java": "No submission",
        "python": "No submission",
        "pascal": "No submission",
        "moss": "100"
    }
]

contest_name = "lmao1"
workbook = Workbook()
workbook.remove(workbook['Sheet'])
lang = ['c','cpp','java','pascal','python']
for problem in data:
    workbook.create_sheet(problem['problem'])
    sheet = workbook[problem['problem']]
    sheet.append(("TÊN NGƯỜI DÙNG","% MOSS"))
    sheet.column_dimensions['A'].width = 20
    moss = problem['moss']
    for check in lang:
        req = get_requests(problem[check]).content if problem[check]!="No submission" else None
        if req:
            soup = BeautifulSoup(req,'html.parser')
            data_tds = soup.find_all('a')[6:]
            for i in data_tds:
                data_split = i.text.split()
                number = ''.join(filter(str.isdigit,data_split[1]))
                sheet.append((data_split[0],number))
    for row in range(2,sheet.max_row+1):
        cell = sheet.cell(row=row,column=2)
        if(int(cell.value)>int(problem['moss'])):
            cell.fill = PatternFill(start_color="ff0d00", end_color="ff0d00", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="34eb43", end_color="34eb43", fill_type="solid")
    
    workbook.save(f"/mnt/moss/{contest_name}.xlsx")
